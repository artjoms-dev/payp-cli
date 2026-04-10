"""Test ExportTool across all 4 formats: csv, json, xlsx, parquet.

Requires a running Postgres container (docker compose up -d postgres)
seeded from tests/seed.sql.

Run directly:

    python tests/test_export_formats.py
"""

from __future__ import annotations

import asyncio
import csv
import json
import sys
import tempfile
import traceback
from pathlib import Path

from payp.db.connection import ConnectionManager
from payp.models import ConnectionCredential, ConnectionProfile, DbType
from payp.tools.export import ExportTool

PROFILE = ConnectionProfile(
    name="pg-local",
    db_type=DbType.POSTGRESQL,
    host="localhost",
    port=5432,
    database="payp_test",
    username="payp",
)
CREDENTIAL = ConnectionCredential(password="payp_dev")

# customers table seed has 30 rows — query them all (and cross-join to get ~100)
SQL = """
    SELECT c.id, c.name, c.email, c.region, c.segment, c.created_at
    FROM customers c
    ORDER BY c.id
    LIMIT 100
"""


async def main() -> int:
    tool = ExportTool()
    conn = ConnectionManager(PROFILE, CREDENTIAL)
    try:
        version = await conn.connect()
        print(f"connected: {version}")
    except Exception as e:
        print(f"FAIL connect: {e}")
        return 1

    context = {"connection_manager": conn}
    tmpdir = Path(tempfile.mkdtemp(prefix="payp_export_"))
    print(f"export dir: {tmpdir}")

    results: dict[str, bool] = {}

    try:
        for fmt in ("csv", "json", "xlsx", "parquet"):
            out = tmpdir / f"customers.{fmt}"
            res = await tool.call(
                {"sql": SQL, "format": fmt, "path": str(out)},
                context,
            )
            if not res.success:
                print(f"  [{fmt}] FAIL: {res.error}")
                results[fmt] = False
                continue
            if not out.exists():
                print(f"  [{fmt}] FAIL: file not created")
                results[fmt] = False
                continue

            size = out.stat().st_size
            rows_written = res.data.get("rows")

            # Validate format-specific readability
            ok = True
            detail = ""
            try:
                if fmt == "csv":
                    with open(out) as f:
                        reader = csv.reader(f)
                        header = next(reader)
                        data_rows = list(reader)
                    detail = f"cols={len(header)} rows={len(data_rows)}"
                    ok = len(data_rows) == rows_written
                elif fmt == "json":
                    # Breaking change: json format now produces JSONL
                    # (newline-delimited JSON) so exports can stream to disk.
                    with open(out) as f:
                        payload = [json.loads(line) for line in f if line.strip()]
                    detail = f"rows={len(payload)} keys={list(payload[0].keys()) if payload else []}"
                    ok = len(payload) == rows_written
                elif fmt == "xlsx":
                    # Streaming xlsx (openpyxl write-only mode) preserves
                    # bold header via WriteOnlyCell but cannot persist
                    # freeze_panes — known openpyxl 3.1.x limitation.
                    from openpyxl import load_workbook
                    wb = load_workbook(out, read_only=False)
                    ws = wb.active
                    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    data_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
                    header_bold = ws["A1"].font.bold
                    detail = f"cols={len(header)} rows={data_count} bold={header_bold}"
                    ok = data_count == rows_written and bool(header_bold)
                    wb.close()
                elif fmt == "parquet":
                    import pyarrow.parquet as pq
                    table = pq.read_table(out)
                    detail = f"cols={len(table.column_names)} rows={table.num_rows} schema={table.schema.names}"
                    ok = table.num_rows == rows_written
            except Exception as e:
                detail = f"read-back error: {e}"
                ok = False

            status = "OK" if ok else "FAIL"
            print(f"  [{fmt}] {status}  {size:>7} B  written={rows_written}  {detail}")
            results[fmt] = ok

    finally:
        await conn.disconnect()

    total_ok = sum(1 for v in results.values() if v)
    print(f"\nresult: {total_ok}/{len(results)} formats passed")
    return 0 if total_ok == len(results) else 1


if __name__ == "__main__":
    try:
        sys.exit(asyncio.run(main()))
    except Exception:
        traceback.print_exc()
        sys.exit(1)
