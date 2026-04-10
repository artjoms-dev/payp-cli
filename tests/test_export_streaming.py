"""Streaming export — unit tests with a mocked ConnectionManager.

These tests do NOT touch a real database. They stub `stream_raw` as an
async generator and verify each format writes the expected number of
rows to disk, that the `.partial` file is cleaned up on success, and
that a failure path cleans up the partial file too.

Run directly:

    python tests/test_export_streaming.py
"""

from __future__ import annotations

import asyncio
import csv
import json
import sys
import tempfile
import traceback
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any

from payp.tools.export import ExportTool


class _FakeConnection:
    """Mimics the ConnectionManager surface used by ExportTool."""

    def __init__(
        self,
        batches: list[list[dict[str, Any]]] | None = None,
        raise_on_batch: int | None = None,
    ) -> None:
        self._batches = batches or []
        self._raise_on_batch = raise_on_batch
        self.is_connected = True

    def stream_raw(
        self,
        sql: str,
        params: tuple | None = None,
        batch_size: int = 10_000,
    ) -> AsyncIterator[list[dict[str, Any]]]:
        raise_on = self._raise_on_batch
        batches = self._batches

        async def _gen() -> AsyncIterator[list[dict[str, Any]]]:
            for i, batch in enumerate(batches):
                if raise_on is not None and i == raise_on:
                    raise RuntimeError("synthetic mid-stream failure")
                yield batch

        return _gen()


def _make_rows(n: int, offset: int = 0) -> list[dict[str, Any]]:
    return [
        {"id": i + offset, "name": f"row_{i + offset}", "value": (i + offset) * 1.5}
        for i in range(n)
    ]


def _chunked(rows: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [rows[i:i + size] for i in range(0, len(rows), size)]


async def _run_export(
    conn: _FakeConnection,
    filepath: Path,
    fmt: str,
) -> dict[str, Any]:
    tool = ExportTool()
    return (
        await tool.call(
            {"sql": "SELECT 1", "format": fmt, "path": str(filepath)},
            {"connection_manager": conn},
        )
    ).__dict__


async def _test_csv(tmpdir: Path) -> tuple[str, bool, str]:
    rows = _make_rows(1000)
    conn = _FakeConnection(_chunked(rows, 250))
    filepath = tmpdir / "out.csv"
    result = await _run_export(conn, filepath, "csv")
    if not result["success"]:
        return ("csv", False, f"tool failed: {result.get('error')}")
    if not filepath.exists():
        return ("csv", False, "file missing")
    if (filepath.with_suffix(".csv.partial")).exists():
        return ("csv", False, "partial file not cleaned up")
    with open(filepath) as f:
        reader = csv.DictReader(f)
        written = list(reader)
    if len(written) != 1000:
        return ("csv", False, f"expected 1000 rows, got {len(written)}")
    if written[0]["name"] != "row_0":
        return ("csv", False, f"first row wrong: {written[0]}")
    if result["data"]["rows"] != 1000:
        return ("csv", False, f"tool reported {result['data']['rows']} rows")
    return ("csv", True, "OK")


async def _test_jsonl(tmpdir: Path) -> tuple[str, bool, str]:
    rows = _make_rows(1000)
    conn = _FakeConnection(_chunked(rows, 250))
    filepath = tmpdir / "out.jsonl"
    result = await _run_export(conn, filepath, "jsonl")
    if not result["success"]:
        return ("jsonl", False, f"tool failed: {result.get('error')}")
    if not filepath.exists():
        return ("jsonl", False, "file missing")
    if (filepath.with_suffix(".jsonl.partial")).exists():
        return ("jsonl", False, "partial file not cleaned up")
    with open(filepath) as f:
        lines = [json.loads(line) for line in f if line.strip()]
    if len(lines) != 1000:
        return ("jsonl", False, f"expected 1000 rows, got {len(lines)}")
    if lines[0]["name"] != "row_0":
        return ("jsonl", False, f"first row wrong: {lines[0]}")
    # Also verify the "json" alias produces JSONL.
    conn2 = _FakeConnection(_chunked(_make_rows(10), 5))
    filepath2 = tmpdir / "alias.json"
    result2 = await _run_export(conn2, filepath2, "json")
    if not result2["success"]:
        return ("jsonl", False, f"json alias failed: {result2.get('error')}")
    with open(filepath2) as f:
        lines2 = [json.loads(line) for line in f if line.strip()]
    if len(lines2) != 10:
        return ("jsonl", False, f"json alias wrote {len(lines2)} rows")
    return ("jsonl", True, "OK")


async def _test_xlsx(tmpdir: Path) -> tuple[str, bool, str]:
    rows = _make_rows(1000)
    conn = _FakeConnection(_chunked(rows, 250))
    filepath = tmpdir / "out.xlsx"
    result = await _run_export(conn, filepath, "xlsx")
    if not result["success"]:
        return ("xlsx", False, f"tool failed: {result.get('error')}")
    if not filepath.exists():
        return ("xlsx", False, "file missing")
    if (filepath.with_suffix(".xlsx.partial")).exists():
        return ("xlsx", False, "partial file not cleaned up")
    # Read back with openpyxl read-only
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    row_count = sum(1 for _ in ws.iter_rows(min_row=2))  # skip header
    wb.close()
    if row_count != 1000:
        return ("xlsx", False, f"expected 1000 data rows, got {row_count}")
    if result["data"]["rows"] != 1000:
        return ("xlsx", False, f"tool reported {result['data']['rows']} rows")
    return ("xlsx", True, "OK")


async def _test_parquet(tmpdir: Path) -> tuple[str, bool, str]:
    rows = _make_rows(1000)
    conn = _FakeConnection(_chunked(rows, 250))
    filepath = tmpdir / "out.parquet"
    result = await _run_export(conn, filepath, "parquet")
    if not result["success"]:
        return ("parquet", False, f"tool failed: {result.get('error')}")
    if not filepath.exists():
        return ("parquet", False, "file missing")
    if (filepath.with_suffix(".parquet.partial")).exists():
        return ("parquet", False, "partial file not cleaned up")
    import pyarrow.parquet as pq
    table = pq.read_table(filepath)
    if table.num_rows != 1000:
        return ("parquet", False, f"expected 1000 rows, got {table.num_rows}")
    if result["data"]["rows"] != 1000:
        return ("parquet", False, f"tool reported {result['data']['rows']} rows")
    return ("parquet", True, "OK")


async def _test_failure_cleanup(tmpdir: Path) -> tuple[str, bool, str]:
    """Mid-stream failure should leave no .partial file behind."""
    rows = _make_rows(500)
    conn = _FakeConnection(_chunked(rows, 100), raise_on_batch=2)
    filepath = tmpdir / "fail.csv"
    result = await _run_export(conn, filepath, "csv")
    if result["success"]:
        return ("failure_cleanup", False, "expected failure, got success")
    if filepath.exists():
        return ("failure_cleanup", False, "final file should not exist on failure")
    partial = filepath.with_suffix(".csv.partial")
    if partial.exists():
        return ("failure_cleanup", False, "partial file not cleaned up after failure")
    return ("failure_cleanup", True, "OK")


async def _test_stream_raw_is_async_gen() -> tuple[str, bool, str]:
    """Signature check — stream_raw must be an async generator function."""
    import inspect
    from payp.db.connection import ConnectionManager
    if not inspect.isasyncgenfunction(ConnectionManager.stream_raw):
        return (
            "stream_raw_sig",
            False,
            "ConnectionManager.stream_raw is not an async generator",
        )
    return ("stream_raw_sig", True, "OK")


async def main() -> int:
    tmpdir = Path(tempfile.mkdtemp(prefix="payp_stream_"))
    print(f"tmp dir: {tmpdir}")

    tests = [
        _test_stream_raw_is_async_gen(),
        _test_csv(tmpdir),
        _test_jsonl(tmpdir),
        _test_xlsx(tmpdir),
        _test_parquet(tmpdir),
        _test_failure_cleanup(tmpdir),
    ]

    results: list[tuple[str, bool, str]] = []
    for coro in tests:
        try:
            results.append(await coro)
        except Exception as e:
            results.append(("exception", False, f"{e}\n{traceback.format_exc()}"))

    fail_count = 0
    for name, ok, msg in results:
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] {name}: {msg}")
        if not ok:
            fail_count += 1

    print()
    if fail_count:
        print(f"✗ {fail_count} test(s) failed")
        return 1
    print(f"✓ all {len(results)} streaming export tests passed")
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
